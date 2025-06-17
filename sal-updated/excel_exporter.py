import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def append_or_create_excel(filename, sheet_name, data, headers, color_hex_list=None, df_measure=None, df_std=None):
    df_new = pd.DataFrame(data, columns=headers)

    if not os.path.isfile(filename):
        # File doesn't exist, so create it and write everything once
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_new.to_excel(writer, index=False, sheet_name=sheet_name)
            if df_measure is not None:
                df_measure.to_excel(writer, sheet_name="Measurement Conditions", index=False)
            if df_std is not None:
                df_std.to_excel(writer, sheet_name="STD value", index=False)
    else:
        # File exists: read and append to avoid duplication
        xls = pd.ExcelFile(filename)
        if sheet_name in xls.sheet_names:
            existing_df = pd.read_excel(xls, sheet_name=sheet_name)
            final_df = pd.concat([existing_df, df_new], ignore_index=True)
        else:
            final_df = df_new

        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            final_df.to_excel(writer, index=False, sheet_name=sheet_name)
            if df_measure is not None:
                df_measure.to_excel(writer, sheet_name="Measurement Conditions", index=False)
            if df_std is not None:
                df_std.to_excel(writer, sheet_name="STD value", index=False)

    # Apply background fill to last row if color provided
    if color_hex_list:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        cl_col = headers.index("Chloride in Soil (mg/kg)") + 1
        start_row = ws.max_row - len(color_hex_list) + 1

        for i, hex_code in enumerate(color_hex_list):
            hex_code = hex_code.replace("#", "").upper()
            if len(hex_code) == 6:
                argb = "FF" + hex_code
            elif len(hex_code) == 8:
                argb = hex_code
            else:
                argb = "FFFFFFFF"

            fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
            ws.cell(row=start_row + i, column=cl_col).fill = fill

        wb.save(filename)
